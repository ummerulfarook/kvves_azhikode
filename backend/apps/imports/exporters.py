"""
Excel export builders using openpyxl.
"""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def style_header_row(ws, row_num=1, fill_color='1A3C5E'):
    """Apply header styling to a row."""
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    font = Font(color='FFFFFF', bold=True)
    for cell in ws[row_num]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center', vertical='center')


def auto_column_width(ws):
    """Auto-size all columns."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def export_members(members_qs):
    """Export members queryset to Excel bytes."""
    wb = Workbook()

    # Sheet 1: Member List
    ws1 = wb.active
    ws1.title = 'Members'
    headers = [
        'Member No', 'Full Name', 'Malayalam Name', 'DOB', 'Gender',
        'Phone', 'Alt Phone', 'Email', 'Address', 'Ward', 'Panchayat',
        'District', 'PIN Code', 'Aadhaar', 'PAN', 'Membership Type',
        'Joining Date', 'Status', 'Remarks',
    ]
    ws1.append(headers)
    style_header_row(ws1)

    for m in members_qs:
        ws1.append([
            m.member_no, m.full_name, m.full_name_ml,
            m.date_of_birth.strftime('%d/%m/%Y') if m.date_of_birth else '',
            m.get_gender_display(), m.phone, m.alternate_phone, m.email,
            m.address, m.ward, m.panchayat, m.district, m.pin_code,
            m.aadhaar_number, m.pan_number, m.get_membership_type_display(),
            m.joining_date.strftime('%d/%m/%Y') if m.joining_date else '',
            m.get_status_display(), m.remarks,
        ])
    auto_column_width(ws1)

    # Sheet 2: Nominees
    ws2 = wb.create_sheet('Nominees')
    ws2.append(['Member No', 'Member Name', 'Nominee Name', 'Relationship', 'Phone', 'Share %', 'Primary'])
    style_header_row(ws2)
    for m in members_qs:
        for nom in m.nominees.all():
            ws2.append([
                m.member_no, m.full_name, nom.name,
                nom.get_relationship_display(), nom.phone,
                str(nom.share_percentage), 'Yes' if nom.is_primary else 'No',
            ])
    auto_column_width(ws2)

    # Sheet 3: Chit Enrollments
    ws3 = wb.create_sheet('Chit Enrollments')
    ws3.append(['Member No', 'Member Name', 'Group No', 'Group Name', 'Ticket No', 'Enrollment Date', 'Status', 'Prize Won'])
    style_header_row(ws3)
    for m in members_qs:
        for enroll in m.chit_enrollments.select_related('chit_group').all():
            ws3.append([
                m.member_no, m.full_name,
                enroll.chit_group.group_no, enroll.chit_group.group_name,
                enroll.ticket_number, str(enroll.enrollment_date),
                enroll.get_status_display(), 'Yes' if enroll.prize_won else 'No',
            ])
    auto_column_width(ws3)

    # Sheet 4: Loans
    ws4 = wb.create_sheet('Loans')
    ws4.append(['Member No', 'Member Name', 'Loan No', 'Loan Type', 'Amount', 'Service Charge (₹)', 'Duration', 'Outstanding', 'Status'])
    style_header_row(ws4)
    for m in members_qs:
        for loan in m.loans.all():
            ws4.append([
                m.member_no, m.full_name, loan.loan_no,
                loan.get_loan_type_display(), str(loan.loan_amount),
                str(loan.service_charge), loan.duration_months,
                str(loan.outstanding_balance), loan.get_status_display(),
            ])
    auto_column_width(ws4)

    # Sheet 5: Deposits
    ws5 = wb.create_sheet('Deposits')
    ws5.append(['Member No', 'Member Name', 'Deposit Type', 'Amount', 'Date', 'Maturity Date', 'Status'])
    style_header_row(ws5)
    for m in members_qs:
        for dep in m.deposits.all():
            ws5.append([
                m.member_no, m.full_name,
                dep.get_deposit_type_display(), str(dep.amount),
                str(dep.deposit_date),
                str(dep.maturity_date) if dep.maturity_date else '',
                dep.get_status_display(),
            ])
    auto_column_width(ws5)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def export_overdue(overdue_list):
    """Export combined overdue list to Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Overdue Payments'
    ws.append(['Type', 'Member No', 'Member Name', 'Amount (₹)', 'Due Date', 'Days Overdue', 'Details'])
    style_header_row(ws)

    for item in overdue_list:
        ws.append([
            item['type'], item['member_no'], item['member_name'],
            str(item['amount']), item['due_date'],
            item['days_overdue'], item.get('detail', ''),
        ])
    auto_column_width(ws)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def get_member_import_template():
    """Generate the member import Excel template."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Members Import Template'
    headers = [
        'member_no', 'full_name', 'full_name_ml', 'date_of_birth',
        'gender', 'phone', 'alternate_phone', 'email', 'address',
        'ward', 'panchayat', 'district', 'pin_code', 'aadhaar_number',
        'pan_number', 'membership_type', 'joining_date', 'masavari_paid_till', 'remarks',
    ]
    ws.append(headers)
    style_header_row(ws)

    # Instructions row
    ws.append([
        'e.g. MEM001', 'Full Name', 'Malayalam Name (optional)',
        'DD/MM/YYYY', 'M/F/O', '10 digits', '10 digits (optional)',
        'email@example.com', 'Full address', 'Ward name',
        'Panchayat name', 'Kannur', '6 digits', '12 digits (optional)',
        'ABCDE1234F (optional)', 'regular/associate/honorary', 'DD/MM/YYYY', 'MM/YYYY or DD/MM/YYYY (optional)', 'Optional remarks',
    ])

    auto_column_width(ws)
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def export_period_report(data):
    """Export period report (income/expense summaries) to Excel bytes."""
    wb = Workbook()

    # Sheet 1: Summary Overview
    ws1 = wb.active
    ws1.title = 'Summary Overview'

    ws1.append(['KVVA Management System — Period Performance Report'])
    ws1.append(['Report Period:', data.get('period', '').capitalize()])
    ws1.append(['Date Range:', f"{data.get('start')} to {data.get('end')}"])
    ws1.append(['Label:', data.get('label', '')])
    ws1.append([])

    ws1.append(['Metric', 'Value / Details'])
    metrics = [
        ('Total Inflow', float(data.get('total_inflow', 0))),
        ('Welfare Collections', float(data.get('welfare_collections', 0))),
        ('Welfare Payments Count', data.get('welfare_count', 0)),
        ('Loan Repayments', float(data.get('loan_repayments', 0))),
        ('Loan Repayments Count', data.get('loan_repayment_count', 0)),
        ('Dues Collected', float(data.get('dues_collected', 0))),
        ('Masavari Collections', float(data.get('masavari_collected', 0))),
        ('Masavari Payments Count', data.get('masavari_count', 0)),
        ('Deposits Made', float(data.get('deposits_made', 0))),
        ('New Members Joined', data.get('new_members', 0)),
        ('New Loans Count', data.get('new_loans', 0)),
        ('New Loans Total Amount', float(data.get('new_loans_amount', 0))),
    ]
    for row in metrics:
        ws1.append(row)

    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 25
    ws1.merge_cells('A1:B1')
    ws1['A1'].font = Font(size=14, bold=True)
    style_header_row(ws1, row_num=6)

    # Sheet 2: New Members List
    new_members = data.get('new_members_list', [])
    if new_members:
        ws2 = wb.create_sheet('New Members')
        ws2.append(['Member No', 'Full Name', 'Joining Date', 'Membership Type', 'Status'])
        style_header_row(ws2)
        for m in new_members:
            ws2.append([
                m.get('member_no'), m.get('full_name'), m.get('joining_date'),
                m.get('membership_type'), m.get('status')
            ])
        auto_column_width(ws2)

    # Sheet 3: Welfare Winners & Payouts List
    welfare_winners = data.get('welfare_winners_list', [])
    if welfare_winners:
        ws_w = wb.create_sheet('Welfare Payouts')
        ws_w.append([
            'Recipient Name', 'Scheme Name', 'Ticket #', 'Prize Amount (₹)', 
            'Surcharge (₹)', 'Late Reduction (₹)', 'Draw Date', 
            'Handover Date', 'Payment Mode', 'Cheque / Ref #'
        ])
        style_header_row(ws_w)
        for w in welfare_winners:
            name = w.get('member__full_name') or w.get('non_member_name')
            member_no = w.get('member__member_no')
            recipient = f"{name} ({member_no})" if member_no else name
            
            # Format payout_payment_mode nicely (e.g. cheque -> Cheque)
            pm = w.get('payout_payment_mode', 'cash')
            pm_display = pm.replace('_', ' ').title() if pm else 'Cash'
            
            ws_w.append([
                recipient,
                w.get('chit_group__group_name'),
                w.get('ticket_number'),
                float(w.get('prize_amount') or 0),
                float(w.get('surcharge_amount') or 0),
                float(w.get('reduction_amount') or 0),
                w.get('prize_date'),
                w.get('received_date') or 'Pending',
                pm_display,
                w.get('cheque_number') or ''
            ])
        auto_column_width(ws_w)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
